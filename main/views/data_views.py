from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from ..models import Shipment, Client
from .core_views import apply_filters, clear_messages
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import datetime
import os
import threading
import time
import shutil
from pathlib import Path
import re


# =============================================================================
# BACKUP SYSTEM
# =============================================================================

# Global variable to track if the backup thread is running
backup_thread = None
last_backup_date = None


def format_file_size(size_in_bytes):
    """Format file size in a human-readable format."""
    if size_in_bytes < 1024:
        return f"{size_in_bytes} B"
    elif size_in_bytes < 1024 * 1024:
        return f"{size_in_bytes / 1024:.1f} KB"
    elif size_in_bytes < 1024 * 1024 * 1024:
        return f"{size_in_bytes / (1024 * 1024):.1f} MB"
    else:
        return f"{size_in_bytes / (1024 * 1024 * 1024):.1f} GB"


def setup_backup_directory():
    """Create backup directories for exports if they don't exist."""
    # Base backup directory
    backup_dir = os.path.join(settings.BASE_DIR, 'backups', 'exports')
    
    # Create subdirectories for each format
    formats = ['excel', 'csv', 'pdf']
    for format_type in formats:
        format_dir = os.path.join(backup_dir, format_type)
        os.makedirs(format_dir, exist_ok=True)
    
    return backup_dir


def weekly_backup_thread():
    """Background thread that runs a backup once per week."""
    global last_backup_date
    
    # Read last backup date from file if it exists
    backup_marker_file = os.path.join(settings.BASE_DIR, 'last_backup.txt')
    if os.path.exists(backup_marker_file):
        with open(backup_marker_file, 'r') as f:
            try:
                last_backup_date = datetime.datetime.fromisoformat(f.read().strip())
            except:
                # If file exists but contains invalid data, default to a week ago
                last_backup_date = datetime.datetime.now() - datetime.timedelta(days=7)
    else:
        # If no file exists, default to a week ago so a backup runs immediately
        last_backup_date = datetime.datetime.now() - datetime.timedelta(days=7)
    
    print(f"Weekly backup thread started. Last backup was on {last_backup_date}")
    
    while True:
        try:
            current_time = datetime.datetime.now()
            
            # Check if a week has passed since the last backup
            if last_backup_date is None or (current_time - last_backup_date).days >= 7:
                print(f"Running weekly backup at {current_time}")
                
                # Create backup - Use optimized queryset
                shipments = Shipment.objects.select_related('client').all()
                timestamp = current_time.strftime("%Y%m%d")
                filename_base = f"weekly_backup_{timestamp}"
                backup_dir = setup_backup_directory()
                
                export_to_excel(shipments, filename_base, backup_dir)
                export_to_csv(shipments, filename_base, backup_dir)
                export_to_pdf(shipments, filename_base, backup_dir)
                
                # Update last backup date
                last_backup_date = current_time
                
                # Save last backup date to file
                with open(backup_marker_file, 'w') as f:
                    f.write(last_backup_date.isoformat())
                
                print(f"Weekly backup completed at {current_time}")
            
            # Sleep for an hour before checking again
            time.sleep(3600)  # 3600 seconds = 1 hour
        except Exception as e:
            print(f"Error in backup thread: {str(e)}")
            # Sleep for a shorter time if there was an error
            time.sleep(300)  # 5 minutes


def start_backup_thread():
    """Start the backup thread if it's not already running."""
    global backup_thread
    if backup_thread is None or not backup_thread.is_alive():
        backup_thread = threading.Thread(target=weekly_backup_thread, daemon=True)
        backup_thread.start()
        print("Weekly backup thread started")
    else:
        print("Weekly backup thread already running")


# Start the backup thread when this module is imported (when Django starts)
start_backup_thread()


# =============================================================================
# EXPORT VIEWS
# =============================================================================

@login_required(login_url='login')
def export_shipments(request):
    """Export shipment data to different formats (Excel, CSV, PDF) and save a backup copy."""
    clear_messages(request)
    
    # Get export format and other parameters
    export_format = request.GET.get('format', 'excel')
    client_id = request.GET.get('client')
    
    # Get shipments with filters if provided - Use optimized queryset
    shipments = Shipment.objects.select_related('client').all()
    shipments = apply_filters(request, shipments)
    
    # Get client name for filename
    client_name = "all_clients"
    if client_id:
        try:
            client = Client.objects.get(pk=client_id)
            client_name = client.name.replace(" ", "_").replace("/", "_")
        except Client.DoesNotExist:
            pass
    
    # Generate timestamp for filename
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create filename based on client and date
    filename_base = f"claims_{client_name}_{timestamp}"
    
    # Ensure backup directory exists
    backup_dir = setup_backup_directory()
    
    # Export based on selected format
    if export_format == 'excel':
        response = export_to_excel(shipments, filename_base, backup_dir)
        return response
    elif export_format == 'csv':
        response = export_to_csv(shipments, filename_base, backup_dir)
        return response
    elif export_format == 'pdf':
        response = export_to_pdf(shipments, filename_base, backup_dir)
        return response
    else:
        messages.error(request, f"Unsupported export format: {export_format}")
        return redirect('shipment_list')


@login_required(login_url='login')
def export_shipments_excel(request):
    """Legacy function that redirects to the more flexible export_shipments function."""
    return export_shipments(request)


# =============================================================================
# EXPORT HELPER FUNCTIONS - MATCHING TABLE COLUMNS EXACTLY
# =============================================================================

def export_to_excel(shipments, filename_base, backup_dir):
    """Helper function to export data to Excel format with local backup - matches table columns exactly."""
    # Create a workbook and active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Shipments'
    
    # Define headers exactly as shown in the table
    headers = [
        'Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 
        'Intent', 'Intent Date', 'Formal', 'Formal Date', 'Value', 
        'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 'Savings',
        'Settlement', 'Exposure', 'Status', 'Closed', 'Actions'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    
    # Add headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    for row_num, shipment in enumerate(shipments, 2):
        # Use the new client-specific shipment numbers
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'
        
        # Format amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"
        
        # Format boolean fields as icons/text
        intent_to_claim = "✓" if shipment.Intent_To_Claim == 'YES' else "✗"
        formal_claim = "✓" if shipment.Formal_Claim_Received == 'YES' else "✗"
        
        # Format status badges
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = '✓ Settled'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = '✗ Not Settled'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = '~ Partial'
        else:
            settlement_status = '-'
        
        status_display = ''
        if shipment.Status == 'OPEN':
            status_display = '● Open'
        elif shipment.Status == 'CLOSED':
            status_display = '✓ Closed'
        elif shipment.Status == 'PENDING':
            status_display = '⏳ Pending'
        elif shipment.Status == 'REJECTED':
            status_display = '✗ Rejected'
        elif shipment.Status == 'UNDER_REVIEW':
            status_display = '◐ Under Review'
        else:
            status_display = shipment.Status
        
        # Create row data matching table exactly
        row_data = [
            shipment.Claim_No,  # This will now be the new format: ClientName-X-YYYYMMDD
            shipment.Brand or '-',
            shipment.Claimant or '-',
            client_id,
            client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_display,
            closed_date,
            'Edit/Delete'  # Actions column placeholder
        ]
        
        # Add row to worksheet
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long text
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save a local backup copy
    excel_backup_path = os.path.join(backup_dir, 'excel', f"{filename_base}.xlsx")
    workbook.save(excel_backup_path)
    
    # Create response for download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    return response


def export_to_csv(shipments, filename_base, backup_dir):
    """Helper function to export data to CSV format with local backup - matches table columns exactly."""
    # Create a file-like buffer for both backup and response
    backup_file = open(os.path.join(backup_dir, 'csv', f"{filename_base}.csv"), 'w', newline='', encoding='utf-8')
    response_buffer = io.StringIO()
    
    # Create CSV writers
    backup_writer = csv.writer(backup_file)
    response_writer = csv.writer(response_buffer)
    
    # Write header row exactly as shown in table
    headers = [
        'Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 
        'Intent', 'Intent Date', 'Formal', 'Formal Date', 'Value', 
        'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 'Savings',
        'Settlement', 'Exposure', 'Status', 'Closed', 'Actions'
    ]
    backup_writer.writerow(headers)
    response_writer.writerow(headers)
    
    # Write data rows
    for shipment in shipments:
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'
        
        # Format amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"
        
        # Format boolean fields
        intent_to_claim = "Yes" if shipment.Intent_To_Claim == 'YES' else "No"
        formal_claim = "Yes" if shipment.Formal_Claim_Received == 'YES' else "No"
        
        # Format status
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = 'Settled'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = 'Not Settled'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = 'Partial'
        else:
            settlement_status = '-'
        
        status_display = shipment.get_Status_display() if shipment.Status else 'Open'
        
        row = [
            shipment.Claim_No,  # New format: ClientName-X-YYYYMMDD
            shipment.Brand or '-',
            shipment.Claimant or '-',
            client_id,
            client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_display,
            closed_date,
            'Edit/Delete'  # Actions column placeholder
        ]
        
        # Write to both files
        backup_writer.writerow(row)
        response_writer.writerow(row)
    
    # Close the backup file
    backup_file.close()
    
    # Create response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
    response.write(response_buffer.getvalue())
    
    return response


def export_to_pdf(shipments, filename_base, backup_dir):
    """Helper function to export data to PDF format with local backup - matches table columns exactly."""
    # Create a file-like buffer for the PDF data
    buffer = io.BytesIO()
    
    # Create the PDF object with landscape orientation for all columns
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        title=f"Claims Report - {filename_base}",
        topMargin=20,
        bottomMargin=20,
        leftMargin=20,
        rightMargin=20
    )
    
    # Create PDF path for backup
    pdf_backup_path = os.path.join(backup_dir, 'pdf', f"{filename_base}.pdf")
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    
    # Add title
    title = Paragraph(f"Claims Report - {datetime.datetime.now().strftime('%Y-%m-%d')}", title_style)
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))  # Add spacing
    
    # Define table data exactly matching the web table
    data = [
        ['Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 'Intent', 'Intent Date', 
         'Formal', 'Formal Date', 'Value', 'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 
         'Savings', 'Settlement', 'Exposure', 'Status', 'Closed']
    ]
    
    # Add shipment data
    for shipment in shipments:
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'
        
        # Format amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"
        
        # Format boolean fields
        intent_to_claim = "✓" if shipment.Intent_To_Claim == 'YES' else "✗"
        formal_claim = "✓" if shipment.Formal_Claim_Received == 'YES' else "✗"
        
        # Format status
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = '✓'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = '✗'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = '~'
        else:
            settlement_status = '-'
        
        status_symbol = ''
        if shipment.Status == 'OPEN':
            status_symbol = '●'
        elif shipment.Status == 'CLOSED':
            status_symbol = '✓'
        elif shipment.Status == 'PENDING':
            status_symbol = '⏳'
        elif shipment.Status == 'REJECTED':
            status_symbol = '✗'
        elif shipment.Status == 'UNDER_REVIEW':
            status_symbol = '◐'
        else:
            status_symbol = shipment.Status[:3] if shipment.Status else 'OPN'
        
        # Truncate shipment number for PDF to fit better
        shipment_no_display = shipment.Claim_No
        if len(shipment.Claim_No) > 15:
            shipment_no_display = shipment.Claim_No[:12] + '...'
        
        row = [
            shipment_no_display,  # Truncated for PDF display
            (shipment.Brand or '-')[:8] + '...' if shipment.Brand and len(shipment.Brand) > 8 else (shipment.Brand or '-'),
            (shipment.Claimant or '-')[:10] + '...' if shipment.Claimant and len(shipment.Claimant) > 10 else (shipment.Claimant or '-'),
            client_id,
            client_name[:12] + '...' if len(client_name) > 12 else client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_symbol,
            closed_date
        ]
        data.append(row)
    
    # Create table with smaller font to fit all columns
    table = Table(data, repeatRows=1)
    
    # Add style to table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),  # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Header text color
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Header alignment
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 8),  # Smaller header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  # Header bottom padding
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Data background
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Data text color
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Data alignment
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Data font
        ('FONTSIZE', (0, 1), (-1, -1), 6),  # Smaller data font size
        ('TOPPADDING', (0, 1), (-1, -1), 2),  # Data top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),  # Data bottom padding
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),  # Grid style
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),  # Box style
        # Alternate row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        # Right align amount columns
        ('ALIGN', (9, 1), (12, -1), 'RIGHT'),  # Value, ISCM, Carrier, Insurance
        ('ALIGN', (14, 1), (14, -1), 'RIGHT'),  # Savings
        ('ALIGN', (16, 1), (16, -1), 'RIGHT'),  # Exposure
    ]))
    
    # Add table to elements
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # Save backup copy
    with open(pdf_backup_path, 'wb') as f:
        f.write(pdf_data)
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    response.write(pdf_data)
    
    return response


# =============================================================================
# IMPORT VIEWS - UPDATED FOR NEW CLIENT-SPECIFIC SHIPMENT IDs
# =============================================================================

@login_required(login_url='login')
def import_shipments(request):
    """Import shipment data from an Excel file, with detailed feedback on the process."""
    clear_messages(request)
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
            return render(request, 'main/import_shipments.html')

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            worksheet = wb.active
            skipped_entries, created_entries, error_entries = process_excel_data(worksheet)
            if created_entries == 0 and not skipped_entries and not error_entries:
                messages.info(request, 'No new entries were created. Check if the data is already up to date.')
            else:
                messages.success(request, f'Successfully created {created_entries} entries. Skipped {len(skipped_entries)} duplicate entries.')
                if error_entries:
                    messages.error(request, f'Errors occurred in {len(error_entries)} entries. {", ".join(error_entries)}')
        except Exception as e:
            messages.error(request, f'An error occurred while processing the file: {str(e)}')

        return render(request, 'main/import_shipments.html')
    return render(request, 'main/import_shipments.html')


def process_excel_data(worksheet):
    """Process Excel data and save valid entries to the database."""
    skipped_entries = []
    created_entries = 0
    error_entries = []
    
    # Updated column order for import:
    # 0: Shipment Number (can be blank - will auto-generate), 1: Brand, 2: Claimant, 
    # 3: Intent To Claim, 4: Intent To Claim Date, 5: Formal Claim, 6: Formal Claim Date, 
    # 7: Value, 8: Paid By ISCM, 9: Paid By Carrier, 10: Paid By Insurance, 11: Branch, 
    # 12: Total Savings, 13: Settled or Not Settled, 14: Financial Exposure, 15: Status
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:  # Skip completely empty rows
            continue
            
        # Try to get claim number from column 0, or leave blank for auto-generation
        claim_no = str(row[0]).strip() if row[0] else ""
        
        # Skip if claim number exists and is already in database
        if claim_no and Shipment.objects.filter(Claim_No=claim_no).exists():
            skipped_entries.append(claim_no)
            continue
            
        try:
            # Get Brand (column 1)
            brand = ""
            if len(row) > 1 and row[1]:
                brand = str(row[1]).strip()
            
            # Get Claimant (column 2) - REQUIRED for client identification
            claimant = ""
            if len(row) > 2 and row[2]:
                claimant = str(row[2]).strip()
            
            # Skip rows without claimant as we need it to identify/create client
            if not claimant:
                error_entries.append(f'Row {worksheet.iter_rows().__next__()}: Missing claimant name')
                continue
            
            # Get or create client based on claimant name
            client, created = Client.objects.get_or_create(
                name__iexact=claimant,
                defaults={'name': claimant}
            )
            
            # Handle date conversions for Intent To Claim Date (column 4)
            intend_date = None
            if len(row) > 4 and row[4]:
                if isinstance(row[4], datetime.date):
                    intend_date = row[4]
                elif isinstance(row[4], str):
                    try:
                        intend_date = datetime.datetime.strptime(row[4], "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            intend_date = datetime.datetime.strptime(row[4], "%d/%m/%Y").date()
                        except ValueError:
                            intend_date = None
            
            # Handle date conversions for Formal Claim Date (column 6)
            formal_date = None
            if len(row) > 6 and row[6]:
                if isinstance(row[6], datetime.date):
                    formal_date = row[6]
                elif isinstance(row[6], str):
                    try:
                        formal_date = datetime.datetime.strptime(row[6], "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            formal_date = datetime.datetime.strptime(row[6], "%d/%m/%Y").date()
                        except ValueError:
                            formal_date = None
            
            # Handle numeric conversions safely
            # Value (column 7)
            claimed_amount = 0
            if len(row) > 7 and row[7]:
                if isinstance(row[7], (int, float)):
                    claimed_amount = float(row[7])
                elif isinstance(row[7], str) and row[7].strip():
                    clean_str = ''.join(c for c in row[7] if c.isdigit() or c == '.')
                    if clean_str:
                        claimed_amount = float(clean_str)
            
            # Paid By ISCM (column 8)
            iscm_amount = 0
            if len(row) > 8 and row[8]:
                if isinstance(row[8], (int, float)):
                    iscm_amount = float(row[8])
                elif isinstance(row[8], str) and row[8].strip():
                    clean_str = ''.join(c for c in row[8] if c.isdigit() or c == '.')
                    if clean_str:
                        iscm_amount = float(clean_str)
            
            # Paid By Carrier (column 9)
            carrier_amount = 0
            if len(row) > 9 and row[9]:
                if isinstance(row[9], (int, float)):
                    carrier_amount = float(row[9])
                elif isinstance(row[9], str) and row[9].strip():
                    clean_str = ''.join(c for c in row[9] if c.isdigit() or c == '.')
                    if clean_str:
                        carrier_amount = float(clean_str)
            
            # Paid By Insurance (column 10)
            insurance_amount = 0
            if len(row) > 10 and row[10]:
                if isinstance(row[10], (int, float)):
                    insurance_amount = float(row[10])
                elif isinstance(row[10], str) and row[10].strip():
                    clean_str = ''.join(c for c in row[10] if c.isdigit() or c == '.')
                    if clean_str:
                        insurance_amount = float(clean_str)
            
            # Total Savings (column 12)
            total_savings = 0
            if len(row) > 12 and row[12]:
                if isinstance(row[12], (int, float)):
                    total_savings = float(row[12])
                elif isinstance(row[12], str) and row[12].strip():
                    clean_str = ''.join(c for c in row[12] if c.isdigit() or c == '.')
                    if clean_str:
                        total_savings = float(clean_str)
            
            # Financial Exposure (column 14)
            financial_exposure = 0
            if len(row) > 14 and row[14]:
                if isinstance(row[14], (int, float)):
                    financial_exposure = float(row[14])
                elif isinstance(row[14], str) and row[14].strip():
                    clean_str = ''.join(c for c in row[14] if c.isdigit() or c == '.')
                    if clean_str:
                        financial_exposure = float(clean_str)
            
            # Convert Intent To Claim to YES/NO format (column 3)
            intent_claim = "NO"
            if len(row) > 3 and row[3]:
                if isinstance(row[3], bool):
                    intent_claim = "YES" if row[3] else "NO"
                elif isinstance(row[3], str):
                    if row[3].upper() in ["YES", "Y", "TRUE", "1"]:
                        intent_claim = "YES"
            
            # Convert Formal Claim to YES/NO format (column 5)
            formal_claim = "NO"
            if len(row) > 5 and row[5]:
                if isinstance(row[5], bool):
                    formal_claim = "YES" if row[5] else "NO"
                elif isinstance(row[5], str):
                    if row[5].upper() in ["YES", "Y", "TRUE", "1"]:
                        formal_claim = "YES"
            
            # Get branch value (column 11)
            branch = ""
            if len(row) > 11 and row[11]:
                branch = str(row[11]).strip()
                # Validate branch code
                if branch and branch not in [choice[0] for choice in Shipment.BRANCH_CHOICES]:
                    branch = ""  # Set to empty if invalid
            
            # Handle Settlement Status (column 13)
            settlement_status = None
            if len(row) > 13 and row[13]:
                settlement_value = str(row[13]).upper().strip()
                if settlement_value in ["SETTLED", "YES", "Y", "TRUE", "1"]:
                    settlement_status = "SETTLED"
                elif settlement_value in ["NOT SETTLED", "NOT_SETTLED", "NO", "N", "FALSE", "0"]:
                    settlement_status = "NOT_SETTLED"
                elif settlement_value in ["PARTIAL", "PARTIALLY SETTLED", "PARTIAL_SETTLED"]:
                    settlement_status = "PARTIAL"
            
            # Handle Status (column 15)
            status = "OPEN"  # Default status
            if len(row) > 15 and row[15]:
                status_value = str(row[15]).upper().strip()
                valid_statuses = [choice[0] for choice in Shipment.STATUS_CHOICES]
                if status_value in valid_statuses:
                    status = status_value
                elif status_value in ["OPEN", "PENDING", "CLOSED", "REJECTED", "UNDER_REVIEW"]:
                    status = status_value
            
            # Create shipment object - Claim_No will be auto-generated if blank
            shipment = Shipment(
                Claim_No=claim_no,  # Leave blank for auto-generation
                client=client,
                Brand=brand,
                Claimant=claimant,
                Intent_To_Claim=intent_claim,
                Intend_Claim_Date=intend_date,
                Formal_Claim_Received=formal_claim,
                Formal_Claim_Date_Received=formal_date,
                Claimed_Amount=claimed_amount,
                Amount_Paid_By_Awa=iscm_amount,
                Amount_Paid_By_Carrier=carrier_amount,
                Amount_Paid_By_Insurance=insurance_amount,
                Branch=branch,
                Total_Savings=total_savings,
                Settlement_Status=settlement_status,
                Financial_Exposure=financial_exposure,
                Status=status
            )
            shipment.save()  # Auto-generates Claim_No if blank
            created_entries += 1
            
        except Exception as e:
            error_entries.append(f'Row with Claimant {claimant}: {str(e)}')

    return skipped_entries, created_entries, error_entries


# =============================================================================
# BACKUP MANAGEMENT VIEWS
# =============================================================================

@login_required(login_url='login')
def browse_backups(request):
    """Browse backup files in the exports directory."""
    clear_messages(request)
    
    # Base path for backups
    base_backup_dir = os.path.join(settings.BASE_DIR, 'backups', 'exports')
    
    # Ensure base backup directory exists
    if not os.path.exists(base_backup_dir):
        os.makedirs(base_backup_dir)
    
    # Subdirectories for different export formats
    backup_formats = ['excel', 'csv', 'pdf']
    
    # Dictionary to store files
    backup_files = {}
    
    # Collect files from each format directory
    for format_type in backup_formats:
        format_dir = os.path.join(base_backup_dir, format_type)
        
        # Create directory if it doesn't exist
        if not os.path.exists(format_dir):
            os.makedirs(format_dir)
            continue
        
        # Collect files and subdirectories
        file_list = []
        try:
            for item in os.listdir(format_dir):
                full_path = os.path.join(format_dir, item)
                
                if os.path.isfile(full_path):
                    # File details
                    file_stat = os.stat(full_path)
                    file_list.append({
                        'name': item,
                        'path': full_path,
                        'size': file_stat.st_size,
                        'size_formatted': format_file_size(file_stat.st_size),
                        'modified': datetime.datetime.fromtimestamp(file_stat.st_mtime)
                    })
        except Exception as e:
            print(f"Error processing files in {format_dir}: {e}")
        
        # Sort files by modification time (newest first)
        file_list.sort(key=lambda x: x['modified'], reverse=True)
        
        # Store files
        backup_files[format_type] = file_list
    
    # We don't pass the actual base_backup_dir to avoid exposing server paths
    return render(request, 'main/browse_backups.html', {
        'backup_files': backup_files,
        'has_backups': bool(sum(len(files) for files in backup_files.values()))
    })


@login_required(login_url='login')
def download_backup(request, format_type, filename):
    """Download a specific backup file."""
    backup_dir = setup_backup_directory()
    file_path = os.path.join(backup_dir, format_type, filename)
    
    if not os.path.exists(file_path):
        messages.error(request, f"File not found: {filename}")
        return redirect('browse_backups')
    
    # Determine content type based on format
    if format_type == 'excel':
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif format_type == 'csv':
        content_type = 'text/csv'
    elif format_type == 'pdf':
        content_type = 'application/pdf'
    else:
        content_type = 'application/octet-stream'
    
    # Create response
    with open(file_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@login_required(login_url='login')
def manual_backup_now(request):
    """Manually trigger a backup."""
    try:
        # Create backup - Use optimized queryset
        current_time = datetime.datetime.now()
        shipments = Shipment.objects.select_related('client').all()
        timestamp = current_time.strftime("%Y%m%d_%H%M%S")
        filename_base = f"manual_backup_{timestamp}"
        backup_dir = setup_backup_directory()
        
        # Create backups in all formats
        export_to_excel(shipments, filename_base, backup_dir)
        export_to_csv(shipments, filename_base, backup_dir)
        export_to_pdf(shipments, filename_base, backup_dir)
        
        messages.success(request, "Manual backup created successfully in all formats.")
    except Exception as e:
        messages.error(request, f"Error creating backup: {str(e)}")
    
    return redirect('browse_backups')



@login_required(login_url='login')
def weekly_backup_status(request):
    """View to show weekly backup status with countdown timer and recent backups."""
    clear_messages(request)
    
    import os
    import datetime
    from django.conf import settings
    
    # Read last backup date from file
    backup_marker_file = os.path.join(settings.BASE_DIR, 'last_backup.txt')
    last_backup_date = None
    next_backup_date = None
    days_until_backup = 0
    hours_until_backup = 0
    backup_overdue = False
    
    if os.path.exists(backup_marker_file):
        try:
            with open(backup_marker_file, 'r') as f:
                last_backup_date = datetime.datetime.fromisoformat(f.read().strip())
                next_backup_date = last_backup_date + datetime.timedelta(days=7)
                
                # Calculate time until next backup
                current_time = datetime.datetime.now()
                time_until = next_backup_date - current_time
                
                if time_until.total_seconds() > 0:
                    days_until_backup = time_until.days
                    hours_until_backup = time_until.seconds // 3600
                else:
                    # Backup is overdue
                    backup_overdue = True
                    overdue_time = current_time - next_backup_date
                    days_until_backup = -overdue_time.days
                    hours_until_backup = -(overdue_time.seconds // 3600)
                    
        except Exception as e:
            print(f"Error reading backup marker file: {e}")
            # Set default values
            last_backup_date = datetime.datetime.now() - datetime.timedelta(days=7)
            next_backup_date = datetime.datetime.now()
            backup_overdue = True
    else:
        # No backup file exists, backup is needed
        backup_overdue = True
        next_backup_date = datetime.datetime.now()
        days_until_backup = 0
        hours_until_backup = 0
    
    # Get recent backup files
    backup_dir = setup_backup_directory()
    recent_backups = []
    
    # Check for weekly backup files in Excel format
    excel_dir = os.path.join(backup_dir, 'excel')
    if os.path.exists(excel_dir):
        try:
            for filename in os.listdir(excel_dir):
                if filename.startswith('weekly_backup_') and filename.endswith('.xlsx'):
                    file_path = os.path.join(excel_dir, filename)
                    file_stat = os.stat(file_path)
                    recent_backups.append({
                        'filename': filename,
                        'size': format_file_size(file_stat.st_size),
                        'created': datetime.datetime.fromtimestamp(file_stat.st_mtime),
                        'format': 'excel'
                    })
        except Exception as e:
            print(f"Error reading backup files: {e}")
    
    # Sort by creation date (newest first) and limit to 10
    recent_backups.sort(key=lambda x: x['created'], reverse=True)
    recent_backups = recent_backups[:10]
    
    # Get total number of claims for backup stats
    total_claims = Shipment.objects.count()
    
    # Manual backup trigger
    if request.method == 'POST' and 'trigger_backup' in request.POST:
        try:
            # Trigger manual backup
            current_time = datetime.datetime.now()
            shipments = Shipment.objects.select_related('client').all()
            timestamp = current_time.strftime("%Y%m%d_%H%M%S")
            filename_base = f"manual_weekly_backup_{timestamp}"
            
            # Create Excel backup
            export_to_excel(shipments, filename_base, backup_dir)
            
            # Update the last backup time
            with open(backup_marker_file, 'w') as f:
                f.write(current_time.isoformat())
            
            messages.success(request, f"Manual weekly backup created successfully! ({total_claims} claims backed up)")
            return redirect('weekly_backup_status')
            
        except Exception as e:
            messages.error(request, f"Error creating manual backup: {str(e)}")
    
    context = {
        'last_backup_date': last_backup_date,
        'next_backup_date': next_backup_date,
        'days_until_backup': abs(days_until_backup),
        'hours_until_backup': abs(hours_until_backup),
        'backup_overdue': backup_overdue,
        'recent_backups': recent_backups,
        'total_claims': total_claims,
        'backup_thread_running': backup_thread is not None and backup_thread.is_alive(),
    }
    
    return render(request, 'main/weekly_backup_status.html', context)


def custom_404(request, exception):
    # Redirect to homepage
    return redirect('index')
