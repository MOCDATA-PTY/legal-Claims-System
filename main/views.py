from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, FileResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import ShipmentForm, LoginForm, RegisterForm, ClientForm
from .models import Shipment, Client
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import datetime
import os
import threading
import time
from django.conf import settings
import shutil
from pathlib import Path
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required
from django.db import transaction

# Global variable to track if the backup thread is running
backup_thread = None
last_backup_date = None

def format_file_size(size_in_bytes):
    """Format file size in a human-readable format."""
    if size_in_bytes < 1024:
        return f"{size_in_bytes} B"
    elif size_in_bytes < 1024 * 1024:
        return f"{size_in_bytes / 1024:.1f} KB"
    elif size_in_bytes < 1024 * 1024 * 1024:
        return f"{size_in_bytes / (1024 * 1024):.1f} MB"
    else:
        return f"{size_in_bytes / (1024 * 1024 * 1024):.1f} GB"

# Function to setup backup directory
def setup_backup_directory():
    """Create backup directories for exports if they don't exist."""
    # Base backup directory
    backup_dir = os.path.join(settings.BASE_DIR, 'backups', 'exports')
    
    # Create subdirectories for each format
    formats = ['excel', 'csv', 'pdf']
    for format_type in formats:
        format_dir = os.path.join(backup_dir, format_type)
        os.makedirs(format_dir, exist_ok=True)
    
    return backup_dir

# Simple thread-based weekly backup system
def weekly_backup_thread():
    """Background thread that runs a backup once per week."""
    global last_backup_date
    
    # Read last backup date from file if it exists
    backup_marker_file = os.path.join(settings.BASE_DIR, 'last_backup.txt')
    if os.path.exists(backup_marker_file):
        with open(backup_marker_file, 'r') as f:
            try:
                last_backup_date = datetime.datetime.fromisoformat(f.read().strip())
            except:
                # If file exists but contains invalid data, default to a week ago
                last_backup_date = datetime.datetime.now() - datetime.timedelta(days=7)
    else:
        # If no file exists, default to a week ago so a backup runs immediately
        last_backup_date = datetime.datetime.now() - datetime.timedelta(days=7)
    
    print(f"Weekly backup thread started. Last backup was on {last_backup_date}")
    
    while True:
        try:
            current_time = datetime.datetime.now()
            
            # Check if a week has passed since the last backup
            if last_backup_date is None or (current_time - last_backup_date).days >= 7:
                print(f"Running weekly backup at {current_time}")
                
                # Create backup
                shipments = Shipment.objects.all()
                timestamp = current_time.strftime("%Y%m%d")
                filename_base = f"weekly_backup_{timestamp}"
                backup_dir = setup_backup_directory()
                
                export_to_excel(shipments, filename_base, backup_dir)
                export_to_csv(shipments, filename_base, backup_dir)
                export_to_pdf(shipments, filename_base, backup_dir)
                
                # Update last backup date
                last_backup_date = current_time
                
                # Save last backup date to file
                with open(backup_marker_file, 'w') as f:
                    f.write(last_backup_date.isoformat())
                
                print(f"Weekly backup completed at {current_time}")
            
            # Sleep for an hour before checking again
            time.sleep(3600)  # 3600 seconds = 1 hour
        except Exception as e:
            print(f"Error in backup thread: {str(e)}")
            # Sleep for a shorter time if there was an error
            time.sleep(300)  # 5 minutes

# Start the backup thread when Django loads
def start_backup_thread():
    """Start the backup thread if it's not already running."""
    global backup_thread
    if backup_thread is None or not backup_thread.is_alive():
        backup_thread = threading.Thread(target=weekly_backup_thread, daemon=True)
        backup_thread.start()
        print("Weekly backup thread started")
    else:
        print("Weekly backup thread already running")

# Start the backup thread when this module is imported (when Django starts)
start_backup_thread()

@login_required(login_url='login')
def clear_database(request):
   if request.method == 'POST':
       with transaction.atomic():
           # Assuming 'Shipment' is the model you want to clear
           Shipment.objects.all().delete()
           messages.success(request, "All shipments have been successfully deleted.")
       return redirect('shipment_list')  # Redirect to a safe page after clearing the database
   else:
       messages.error(request, "Invalid request method.")
       return redirect('shipment_list')


def clear_messages(request):
   """Helper function to clear all messages from the request."""
   storage = messages.get_messages(request)
   for message in storage:
       pass
   storage.used = True

def index(request):
   """Redirect root URL to login page."""
   return redirect('login')

@login_required(login_url='login')
def home(request):
   """Home page view that requires login."""
   clear_messages(request)
   return render(request, 'main/home.html')

# Client management views
@login_required(login_url='login')
def client_list(request):
   """View to list all clients."""
   clear_messages(request)
   clients = Client.objects.all().order_by('name')
   return render(request, 'main/client_list.html', {'clients': clients})

@login_required(login_url='login')
def add_client(request):
   """Add a new client."""
   clear_messages(request)
   if request.method == 'POST':
       form = ClientForm(request.POST)
       if form.is_valid():
           client = form.save()
           messages.success(request, f'Client added successfully with ID: {client.client_id}')
           return redirect('client_list')
       else:
           messages.error(request, 'Please correct the errors below.')
   else:
       form = ClientForm()
   
   return render(request, 'main/add_client.html', {'form': form})

@login_required(login_url='login')
def edit_client(request, pk):
   """Edit an existing client."""
   clear_messages(request)
   client = get_object_or_404(Client, pk=pk)
   
   if request.method == 'POST':
       form = ClientForm(request.POST, instance=client)
       if form.is_valid():
           form.save()
           messages.success(request, 'Client updated successfully.')
           return redirect('client_list')
       else:
           messages.error(request, 'Please correct the errors below.')
   else:
       form = ClientForm(instance=client)
   
   return render(request, 'main/edit_client.html', {'form': form, 'client': client})

@login_required(login_url='login')
def delete_client(request, pk):
   """Delete a client."""
   clear_messages(request)
   client = get_object_or_404(Client, pk=pk)
   
   # Check if there are shipments associated with this client
   shipment_count = Shipment.objects.filter(client=client).count()
   
   if request.method == 'POST':
       if shipment_count > 0 and 'confirm' not in request.POST:
           messages.warning(request, f'This client has {shipment_count} shipments. Are you sure you want to delete?')
           return render(request, 'main/delete_client_confirm.html', {'client': client, 'shipment_count': shipment_count})
       
       # Delete client and any associated shipments
       client.delete()
       messages.success(request, 'Client deleted successfully.')
       return redirect('client_list')
   
   return render(request, 'main/delete_client_confirm.html', {'client': client, 'shipment_count': shipment_count})

@login_required(login_url='login')
def add_shipment(request):
   """Add a new shipment."""
   # Get all clients for dropdown
   clients = Client.objects.all().order_by('name')

   if request.method == 'POST':
       form = ShipmentForm(request.POST)
       claim_no = request.POST.get('Claim_No')

       if Shipment.objects.filter(Claim_No=claim_no).exists():
           existing_shipment = Shipment.objects.get(Claim_No=claim_no)
           messages.warning(request, f'Duplicate claim number {claim_no}, consider editing the existing entry.')
           return render(request, 'main/add_shipment.html', {
               'form': form,
               'clients': clients,
               'duplicate_claim_no': claim_no,
               'edit_shipment_id': existing_shipment.id
           })

       if form.is_valid():
           # Check if client exists
           client_name = form.cleaned_data.get('client_name')
           client, created = Client.objects.get_or_create(
               name__iexact=client_name,
               defaults={'name': client_name}
           )
           
           # Add success message about client
           if created:
               messages.info(request, f'New client "{client_name}" created with ID: {client.client_id}')
           else:
               messages.info(request, f'Using existing client "{client_name}" (ID: {client.client_id})')
           
           # Save the form
           form.save()
           messages.success(request, 'Shipment added successfully.')
           return redirect('shipment_list')
       else:
           messages.error(request, 'Please correct the errors below.')

   else:
       form = ShipmentForm()

   return render(request, 'main/add_shipment.html', {
       'form': form,
       'clients': clients
   })

@login_required(login_url='login')
def shipment_list(request):
   """List all shipments with the option to apply filters."""
   clear_messages(request)
   shipments = Shipment.objects.all()
   branches = Shipment.objects.values_list('Branch', flat=True).distinct().order_by('Branch')
   clients = Client.objects.all().order_by('name')
   
   # Apply filters
   shipments = apply_filters(request, shipments)
   
   # Check if we need to handle legacy data
   for shipment in shipments:
       if not hasattr(shipment, 'client') or shipment.client is None:
           # Create a client for this shipment if it doesn't have one
           if hasattr(shipment, 'Claiming_Client') and shipment.Claiming_Client:
               client, created = Client.objects.get_or_create(
                   name=shipment.Claiming_Client,
                   defaults={'name': shipment.Claiming_Client}
               )
               # Link the client to the shipment
               shipment.client = client
               shipment.save(update_fields=['client'])
   
   return render(request, 'main/shipment_list.html', {
       'shipments': shipments,
       'branches': branches,
       'clients': clients,
   })

@login_required(login_url='login')
def edit_shipment(request, pk):
   """Edit an existing shipment."""
   clear_messages(request)
   shipment = get_object_or_404(Shipment, pk=pk)
   clients = Client.objects.all().order_by('name')
   
   # Check if this is a cancel/keep original request
   if request.method == 'POST' and 'keep_original' in request.POST:
       messages.info(request, 'No changes were made to the shipment.')
       return redirect('shipment_list')
   
   if request.method == 'POST':
       form = ShipmentForm(request.POST, instance=shipment)
       if form.is_valid():
           # Check if client exists
           client_name = form.cleaned_data.get('client_name')
           client, created = Client.objects.get_or_create(
               name__iexact=client_name,
               defaults={'name': client_name}
           )
           
           # Add success message about client
           if created:
               messages.info(request, f'New client "{client_name}" created with ID: {client.client_id}')
           else:
               messages.info(request, f'Using existing client "{client_name}" (ID: {client.client_id})')
           
           # Save the form
           form.save()
           messages.success(request, 'Shipment updated successfully.')
           return redirect('shipment_list')
       else:
           messages.error(request, 'Please correct the errors below.')
   else:
       form = ShipmentForm(instance=shipment)
   
   return render(request, 'main/edit_shipment.html', {
       'form': form, 
       'shipment': shipment,
       'clients': clients
   })

@login_required(login_url='login')
def export_shipments(request):
    """Export shipment data to different formats (Excel, CSV, PDF) and save a backup copy."""
    clear_messages(request)
    
    # Get export format and other parameters
    export_format = request.GET.get('format', 'excel')
    client_id = request.GET.get('client')
    
    # Get shipments with filters if provided
    shipments = Shipment.objects.all()
    shipments = apply_filters(request, shipments)
    
    # Get client name for filename
    client_name = "all_clients"
    if client_id:
        try:
            client = Client.objects.get(pk=client_id)
            client_name = client.name.replace(" ", "_").replace("/", "_")
        except Client.DoesNotExist:
            pass
    
    # Generate timestamp for filename
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create filename based on client and date
    filename_base = f"claims_{client_name}_{timestamp}"
    
    # Ensure backup directory exists
    backup_dir = setup_backup_directory()
    
    # Export based on selected format
    if export_format == 'excel':
        response = export_to_excel(shipments, filename_base, backup_dir)
        return response
    elif export_format == 'csv':
        response = export_to_csv(shipments, filename_base, backup_dir)
        return response
    elif export_format == 'pdf':
        response = export_to_pdf(shipments, filename_base, backup_dir)
        return response
    else:
        messages.error(request, f"Unsupported export format: {export_format}")
        return redirect('shipment_list')

def export_to_excel(shipments, filename_base, backup_dir):
    """Helper function to export data to Excel format with local backup."""
    # Create a workbook and active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Shipments'
    
    # Define headers
    headers = [
        'Claim No', 'Claim ID', 'Client Name', 'Branch', 'Formal Claim', 
        'Intend Date', 'Formal Date', 'Claimed Amount', 'Amount Paid by Carrier', 
        'Amount Paid by AWA', 'Amount Paid by Insurance', 'Closed Date'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    
    # Add headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    for row_num, shipment in enumerate(shipments, 2):
        client_id = shipment.client.client_id if hasattr(shipment, 'client') and shipment.client else 'N/A'
        client_name = shipment.client.name if hasattr(shipment, 'client') and shipment.client else (
            shipment.Claiming_Client if hasattr(shipment, 'Claiming_Client') else 'Unknown')
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%Y-%m-%d") if shipment.Intend_Claim_Date else ''
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%Y-%m-%d") if shipment.Formal_Claim_Date_Received else ''
        closed_date = shipment.Closed_Date.strftime("%Y-%m-%d") if shipment.Closed_Date else ''
        
        # Create row data
        row_data = [
            shipment.Claim_No,
            client_id,
            client_name,
            shipment.Branch,
            shipment.Formal_Claim_Received,
            intend_date,
            formal_date,
            shipment.Claimed_Amount,
            shipment.Amount_Paid_By_Carrier,
            shipment.Amount_Paid_By_Awa,
            shipment.Amount_Paid_By_Insurance,
            closed_date
        ]
        
        # Add row to worksheet
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save a local backup copy
    excel_backup_path = os.path.join(backup_dir, 'excel', f"{filename_base}.xlsx")
    workbook.save(excel_backup_path)
    
    # Create response for download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    return response

def export_to_csv(shipments, filename_base, backup_dir):
    """Helper function to export data to CSV format with local backup."""
    # Create a file-like buffer for both backup and response
    backup_file = open(os.path.join(backup_dir, 'csv', f"{filename_base}.csv"), 'w', newline='')
    response_buffer = io.StringIO()
    
    # Create CSV writers
    backup_writer = csv.writer(backup_file)
    response_writer = csv.writer(response_buffer)
    
    # Write header row
    headers = [
        'Claim No', 'Claim ID', 'Client Name', 'Branch', 'Formal Claim', 
        'Intend Date', 'Formal Date', 'Claimed Amount', 'Amount Paid by Carrier', 
        'Amount Paid by AWA', 'Amount Paid by Insurance', 'Closed Date'
    ]
    backup_writer.writerow(headers)
    response_writer.writerow(headers)
    
    # Write data rows
    for shipment in shipments:
        client_id = shipment.client.client_id if hasattr(shipment, 'client') and shipment.client else 'N/A'
        client_name = shipment.client.name if hasattr(shipment, 'client') and shipment.client else (
            shipment.Claiming_Client if hasattr(shipment, 'Claiming_Client') else 'Unknown')
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%Y-%m-%d") if shipment.Intend_Claim_Date else ''
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%Y-%m-%d") if shipment.Formal_Claim_Date_Received else ''
        closed_date = shipment.Closed_Date.strftime("%Y-%m-%d") if shipment.Closed_Date else ''
        
        row = [
            shipment.Claim_No,
            client_id,
            client_name,
            shipment.Branch,
            shipment.Formal_Claim_Received,
            intend_date,
            formal_date,
            shipment.Claimed_Amount,
            shipment.Amount_Paid_By_Carrier,
            shipment.Amount_Paid_By_Awa,
            shipment.Amount_Paid_By_Insurance,
            closed_date
        ]
        
        # Write to both files
        backup_writer.writerow(row)
        response_writer.writerow(row)
    
    # Close the backup file
    backup_file.close()
    
    # Create response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
    response.write(response_buffer.getvalue())
    
    return response

def export_to_pdf(shipments, filename_base, backup_dir):
    """Helper function to export data to PDF format with local backup."""
    # Create a file-like buffer for the PDF data
    buffer = io.BytesIO()
    
    # Create the PDF object
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        title=f"Claims Report - {filename_base}"
    )
    
    # Create PDF path for backup
    pdf_backup_path = os.path.join(backup_dir, 'pdf', f"{filename_base}.pdf")
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    
    # Add title
    title = Paragraph(f"Claims Report - {datetime.datetime.now().strftime('%Y-%m-%d')}", title_style)
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))  # Add spacing
    
    # Define table data
    data = [
        ['Claim No', 'Claim ID', 'Client Name', 'Branch', 'Formal', 'Intend Date', 
         'Formal Date', 'Claimed Amount', 'Carrier Paid', 'AWA Paid', 'Insurance', 'Closed Date']
    ]
    
    # Add shipment data
    for shipment in shipments:
        client_id = shipment.client.client_id if hasattr(shipment, 'client') and shipment.client else 'N/A'
        client_name = shipment.client.name if hasattr(shipment, 'client') and shipment.client else (
            shipment.Claiming_Client if hasattr(shipment, 'Claiming_Client') else 'Unknown')
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%Y-%m-%d") if shipment.Intend_Claim_Date else ''
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%Y-%m-%d") if shipment.Formal_Claim_Date_Received else ''
        closed_date = shipment.Closed_Date.strftime("%Y-%m-%d") if shipment.Closed_Date else ''
        
        # Format amounts
        claimed = f"${shipment.Claimed_Amount:.2f}" if shipment.Claimed_Amount else '$0.00'
        carrier = f"${shipment.Amount_Paid_By_Carrier:.2f}" if shipment.Amount_Paid_By_Carrier else '$0.00'
        awa = f"${shipment.Amount_Paid_By_Awa:.2f}" if shipment.Amount_Paid_By_Awa else '$0.00'
        insurance = f"${shipment.Amount_Paid_By_Insurance:.2f}" if shipment.Amount_Paid_By_Insurance else '$0.00'
        
        row = [
            shipment.Claim_No,
            client_id,
            client_name[:20] + '...' if len(client_name) > 23 else client_name,  # Truncate long names
            shipment.Branch,
            'Yes' if shipment.Formal_Claim_Received == 'YES' else 'No',
            intend_date,
            formal_date,
            claimed,
            carrier,
            awa,
            insurance,
            closed_date
        ]
        data.append(row)
    
    # Create table
    table = Table(data, repeatRows=1)
    
    # Add style to table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),  # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Header text color
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Header alignment
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 10),  # Header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Header bottom padding
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Data background
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Data text color
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Data alignment
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Data font
        ('FONTSIZE', (0, 1), (-1, -1), 8),  # Data font size
        ('TOPPADDING', (0, 1), (-1, -1), 4),  # Data top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),  # Data bottom padding
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),  # Grid style
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),  # Box style
        # Alternate row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
    ]))
    
    # Add table to elements
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # Save backup copy
    with open(pdf_backup_path, 'wb') as f:
        f.write(pdf_data)
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    response.write(pdf_data)
    
    return response

@login_required(login_url='login')
def browse_backups(request):
    """Browse backup files in the exports directory."""
    clear_messages(request)
    
    # Base path for backups
    base_backup_dir = os.path.join(settings.BASE_DIR, 'backups', 'exports')
    
    # Ensure base backup directory exists
    if not os.path.exists(base_backup_dir):
        os.makedirs(base_backup_dir)
    
    # Subdirectories for different export formats
    backup_formats = ['excel', 'csv', 'pdf']
    
    # Dictionary to store files
    backup_files = {}
    
    # Collect files from each format directory
    for format_type in backup_formats:
        format_dir = os.path.join(base_backup_dir, format_type)
        
        # Create directory if it doesn't exist
        if not os.path.exists(format_dir):
            os.makedirs(format_dir)
            continue
        
        # Collect files and subdirectories
        file_list = []
        try:
            for item in os.listdir(format_dir):
                full_path = os.path.join(format_dir, item)
                
                if os.path.isfile(full_path):
                    # File details
                    file_stat = os.stat(full_path)
                    file_list.append({
                        'name': item,
                        'path': full_path,
                        'size': file_stat.st_size,
                        'size_formatted': format_file_size(file_stat.st_size),
                        'modified': datetime.datetime.fromtimestamp(file_stat.st_mtime)
                    })
        except Exception as e:
            print(f"Error processing files in {format_dir}: {e}")
        
        # Sort files by modification time (newest first)
        file_list.sort(key=lambda x: x['modified'], reverse=True)
        
        # Store files
        backup_files[format_type] = file_list
    
    # We don't pass the actual base_backup_dir to avoid exposing server paths
    return render(request, 'main/browse_backups.html', {
        'backup_files': backup_files,
        'has_backups': bool(sum(len(files) for files in backup_files.values()))
    })

@login_required(login_url='login')
def download_backup(request, format_type, filename):
    """Download a specific backup file."""
    backup_dir = setup_backup_directory()
    file_path = os.path.join(backup_dir, format_type, filename)
    
    if not os.path.exists(file_path):
        messages.error(request, f"File not found: {filename}")
        return redirect('browse_backups')
    
    # Determine content type based on format
    if format_type == 'excel':
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif format_type == 'csv':
        content_type = 'text/csv'
    elif format_type == 'pdf':
        content_type = 'application/pdf'
    else:
        content_type = 'application/octet-stream'
    
    # Create response
    with open(file_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

@login_required(login_url='login')
def manual_backup_now(request):
    """Manually trigger a backup."""
    try:
        # Create backup
        current_time = datetime.datetime.now()
        shipments = Shipment.objects.all()
        timestamp = current_time.strftime("%Y%m%d_%H%M%S")
        filename_base = f"manual_backup_{timestamp}"
        backup_dir = setup_backup_directory()
        
        # Create backups in all formats
        export_to_excel(shipments, filename_base, backup_dir)
        export_to_csv(shipments, filename_base, backup_dir)
        export_to_pdf(shipments, filename_base, backup_dir)
        
        messages.success(request, "Manual backup created successfully in all formats.")
    except Exception as e:
        messages.error(request, f"Error creating backup: {str(e)}")
    
    return redirect('browse_backups')

# Legacy function for compatibility
@login_required(login_url='login')
def export_shipments_excel(request):
    """Legacy function that redirects to the more flexible export_shipments function."""
    return export_shipments(request)

@login_required(login_url='login')
def import_shipments(request):
   """Import shipment data from an Excel file, with detailed feedback on the process."""
   clear_messages(request)
   if request.method == 'POST' and request.FILES.get('excel_file'):
       excel_file = request.FILES['excel_file']
       if not excel_file.name.endswith(('.xlsx', '.xls')):
           messages.error(request, 'Invalid file format. Please upload an Excel file.')
           return render(request, 'main/import_shipments.html')

       try:
           wb = openpyxl.load_workbook(excel_file, data_only=True)
           worksheet = wb.active
           skipped_entries, created_entries, error_entries = process_excel_data(worksheet)
           if created_entries == 0 and not skipped_entries and not error_entries:
               messages.info(request, 'No new entries were created. Check if the data is already up to date.')
           else:
               messages.success(request, f'Successfully created {created_entries} entries. Skipped {len(skipped_entries)} duplicate entries.')
               if error_entries:
                   messages.error(request, f'Errors occurred in {len(error_entries)} entries. {", ".join(error_entries)}')
       except Exception as e:
           messages.error(request, f'An error occurred while processing the file: {str(e)}')

       return render(request, 'main/import_shipments.html')
   return render(request, 'main/import_shipments.html')

@login_required(login_url='login')
def client_autocomplete(request):
   """Autocomplete for client names."""
   term = request.GET.get('term', '')
   clients = Client.objects.filter(name__icontains=term).values('id', 'name', 'client_id')
   
   suggestions = [{'id': client['id'], 'text': f"{client['name']} ({client['client_id']})"} for client in clients]

   return JsonResponse(suggestions, safe=False)

def user_login(request):
   if request.method == 'POST':
       form = LoginForm(request, data=request.POST)
       username = request.POST.get('username')
       password = request.POST.get('password')
       user = authenticate(username=username, password=password)
       if user:
           if user.is_active:
               login(request, user)
               request.session.set_expiry(0)  # Expire session on browser close
               return redirect('home')
       else:
           messages.error(request, "Invalid username or password. Please try again.")
   else:
       form = LoginForm()

   response = render(request, 'main/login.html', {'form': form})
   response['Cache-Control'] = 'no-store'
   return response

def register(request):
   """Handle new user registration and automatic login upon successful registration."""
   clear_messages(request)
   if request.method == 'POST':
       form = RegisterForm(request.POST)
       if form.is_valid():
           user = form.save()
           login(request, user)
           request.session.set_expiry(0)
           return redirect('home')
       else:
           messages.error(request, "Error in form submission.")
   else:
       form = RegisterForm()
   return render(request, 'main/register.html', {'form': form})

def user_logout(request):
   """Handle user logout and redirect to login page."""
   clear_messages(request)
   logout(request)  # Clears user session
   return redirect('login')

@login_required(login_url='login')
def delete_shipment(request, pk):
   """Delete a shipment entry."""
   clear_messages(request)
   shipment = get_object_or_404(Shipment, pk=pk)
   if request.method == 'POST':
       shipment.delete()
       messages.success(request, 'Shipment deleted successfully.')
   return redirect('shipment_list')

# Helper functions below
def apply_filters(request, shipments):
   """Apply filters to the shipments queryset based on request parameters."""
   
   # Filter by claim number
   claim_no = request.GET.get('claim_no')
   if claim_no:
       shipments = shipments.filter(Claim_No__icontains=claim_no)
   
   # Filter by client (using the client id or name)
   client_id = request.GET.get('client')
   if client_id:
       try:
           # First try if client_id is numeric (direct ID)
           client_id = int(client_id)
           shipments = shipments.filter(client_id=client_id)
       except ValueError:
           # If not numeric, search by name
           shipments = shipments.filter(client__name__icontains=client_id)
   
   # Filter by client ID (specific Client ID field)
   client_unique_id = request.GET.get('client_unique_id')
   if client_unique_id:
       shipments = shipments.filter(client__client_id=client_unique_id)
   
   # Filter by branch
   branch = request.GET.get('branch')
   if branch:
       shipments = shipments.filter(Branch=branch)
   
   # Filter by Intend Date range
   intend_date_from = request.GET.get('intend_date_from')
   if intend_date_from:
       shipments = shipments.filter(Intend_Claim_Date__gte=intend_date_from)
   
   intend_date_to = request.GET.get('intend_date_to')
   if intend_date_to:
       shipments = shipments.filter(Intend_Claim_Date__lte=intend_date_to)
   
   # Filter by Formal Date range
   formal_date_from = request.GET.get('formal_date_from')
   if formal_date_from:
       shipments = shipments.filter(Formal_Claim_Date_Received__gte=formal_date_from)
   
   formal_date_to = request.GET.get('formal_date_to')
   if formal_date_to:
       shipments = shipments.filter(Formal_Claim_Date_Received__lte=formal_date_to)
   
   return shipments

def process_excel_data(worksheet):
    """Process Excel data and save valid entries to the database."""
    skipped_entries = []
    created_entries = 0
    error_entries = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:  # Skip empty rows or rows without Claim No
            continue
        claim_no = str(row[0])
        if Shipment.objects.filter(Claim_No=claim_no).exists():
            skipped_entries.append(claim_no)
            continue
        try:
            # Handle date conversions
            intend_date = None
            if row[4]:
                if isinstance(row[4], datetime.date):
                    intend_date = row[4]
                elif isinstance(row[4], str):
                    try:
                        intend_date = datetime.datetime.strptime(row[4], "%Y-%m-%d").date()
                    except ValueError:
                        # Try other date formats if the first one fails
                        try:
                            intend_date = datetime.datetime.strptime(row[4], "%d/%m/%Y").date()
                        except ValueError:
                            intend_date = None
            
            formal_date = None
            if row[5]:
                if isinstance(row[5], datetime.date):
                    formal_date = row[5]
                elif isinstance(row[5], str):
                    try:
                        formal_date = datetime.datetime.strptime(row[5], "%Y-%m-%d").date()
                    except ValueError:
                        # Try other date formats if the first one fails
                        try:
                            formal_date = datetime.datetime.strptime(row[5], "%d/%m/%Y").date()
                        except ValueError:
                            formal_date = None
            
            closed_date = None
            if row[10]:
                if isinstance(row[10], datetime.date):
                    closed_date = row[10]
                elif isinstance(row[10], str):
                    try:
                        closed_date = datetime.datetime.strptime(row[10], "%Y-%m-%d").date()
                    except ValueError:
                        # Try other date formats if the first one fails
                        try:
                            closed_date = datetime.datetime.strptime(row[10], "%d/%m/%Y").date()
                        except ValueError:
                            closed_date = None
            
            # Handle numeric conversions safely
            claimed_amount = 0
            if row[6]:
                if isinstance(row[6], (int, float)):
                    claimed_amount = float(row[6])
                elif isinstance(row[6], str) and row[6].strip():
                    # Remove any non-numeric characters except decimal point
                    clean_str = ''.join(c for c in row[6] if c.isdigit() or c == '.')
                    if clean_str:
                        claimed_amount = float(clean_str)
            
            carrier_amount = 0
            if row[7]:
                if isinstance(row[7], (int, float)):
                    carrier_amount = float(row[7])
                elif isinstance(row[7], str) and row[7].strip():
                    clean_str = ''.join(c for c in row[7] if c.isdigit() or c == '.')
                    if clean_str:
                        carrier_amount = float(clean_str)
            
            awa_amount = 0
            if row[8]:
                if isinstance(row[8], (int, float)):
                    awa_amount = float(row[8])
                elif isinstance(row[8], str) and row[8].strip():
                    clean_str = ''.join(c for c in row[8] if c.isdigit() or c == '.')
                    if clean_str:
                        awa_amount = float(clean_str)
            
            insurance_amount = 0
            if row[9]:
                if isinstance(row[9], (int, float)):
                    insurance_amount = float(row[9])
                elif isinstance(row[9], str) and row[9].strip():
                    clean_str = ''.join(c for c in row[9] if c.isdigit() or c == '.')
                    if clean_str:
                        insurance_amount = float(clean_str)
            
            # Convert formal claim received to YES/NO format
            formal_claim = "NO"
            if row[3]:
                if isinstance(row[3], bool):
                    formal_claim = "YES" if row[3] else "NO"
                elif isinstance(row[3], str):
                    if row[3].upper() in ["YES", "Y", "TRUE", "1"]:
                        formal_claim = "YES"
            
            # Get branch value and validate it's in BRANCH_CHOICES
            branch = row[2] or ""
            # Ensure branch code is valid (you may want to add more validation)
            if branch and branch not in [choice[0] for choice in Shipment.BRANCH_CHOICES]:
                branch = ""  # Set to empty if invalid
            
            # Get or create client
            client_name = row[1] or "Unknown Client"
            client, created = Client.objects.get_or_create(
                name__iexact=client_name,
                defaults={'name': client_name}
            )
            
            shipment = Shipment(
                Claim_No=claim_no,
                client=client,
                Branch=branch,
                Formal_Claim_Received=formal_claim,
                Intend_Claim_Date=intend_date,
                Formal_Claim_Date_Received=formal_date,
                Claimed_Amount=claimed_amount,
                Amount_Paid_By_Carrier=carrier_amount,
                Amount_Paid_By_Awa=awa_amount,
                Amount_Paid_By_Insurance=insurance_amount,
                Closed_Date=closed_date
            )
            shipment.save()
            created_entries += 1
        except Exception as e:
            error_entries.append(f'Row with Claim No {claim_no}: {str(e)}')

    return skipped_entries, created_entries, error_entries